import uuid
import io
import os
import csv
import time
import hashlib
from datetime import datetime, timezone
from fastapi import UploadFile, HTTPException, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
import openpyxl

from app.repositories.calificaciones import CalificacionesRepository
from app.repositories.umbral_materia import UmbralMateriaRepository
from app.models.padron import EntradaPadron
from app.models.user import User
from app.services.audit import AuditService


_preview_store: dict[str, dict] = {}

METADATA_COLUMNS = {"nombre", "apellidos", "email", "comision", "regional", "dni", "legajo"}
TEXTUAL_VALUES = {
    "Satisfactorio", "Supera lo esperado", "No satisfactorio",
    "Excelente", "Regular", "Bien", "Muy Bien", "Aprobado", "Reprobado",
}
NUMERIC_SUFFIX = "(Real)"
ALLOWED_EXTENSIONS = {".xlsx", ".csv"}
MAX_FILE_SIZE = 10 * 1024 * 1024
PREVIEW_TTL = 900  # 15 minutes
DEFAULT_UMBRAL_PCT = 60
DEFAULT_VALORES_APROBATORIOS = ["Satisfactorio", "Supera lo esperado"]


def _clean_actividad_name(header: str) -> str:
    result = header
    for suffix in (f" {NUMERIC_SUFFIX}", NUMERIC_SUFFIX, f" {NUMERIC_SUFFIX.lower()}", NUMERIC_SUFFIX.lower()):
        if result.lower().endswith(suffix.lower()):
            result = result[:-len(suffix)]
            break
    return result.strip()


def _is_metadata_column(header: str) -> bool:
    return header.strip().lower() in METADATA_COLUMNS


def _classify_column(header: str, sample_values: list) -> dict | None:
    cleaned = header.strip()
    if _is_metadata_column(cleaned):
        return None

    cleaned_lower = cleaned.lower()
    suffix_lower = NUMERIC_SUFFIX.lower()
    if cleaned_lower.endswith(suffix_lower) or cleaned_lower.endswith(f" {suffix_lower}"):
        numeric_values = [v for v in sample_values if v is not None and isinstance(v, (int, float))]
        return {
            "nombre": _clean_actividad_name(cleaned),
            "tipo": "numerica",
            "valores_muestra": numeric_values[:5],
        }

    text_values = [str(v).strip() for v in sample_values if v is not None and str(v).strip()]
    if any(v in TEXTUAL_VALUES for v in text_values):
        return {
            "nombre": cleaned,
            "tipo": "textual",
            "valores_muestra": list(set(text_values))[:5],
        }

    return None


def _detect_csv_delimiter(head: str) -> str:
    comma = head.count(",")
    semicolon = head.count(";")
    return ";" if semicolon > comma else ","


def _parse_xlsx(content: bytes) -> tuple[list[dict], list[dict], list[str]]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return [], [], ["Archivo vacío"]

    raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    activity_columns = []
    entries = []
    errors = []

    for col_idx in range(len(raw_headers)):
        sample_values = []
        for row in rows[1:11]:  # sample first 10 data rows
            if col_idx < len(row) and row[col_idx] is not None:
                sample_values.append(row[col_idx])
        col_info = _classify_column(raw_headers[col_idx], sample_values)
        if col_info:
            activity_columns.append({**col_info, "col_idx": col_idx})

    for row_idx, row in enumerate(rows[1:], start=2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        email_val = ""
        nombre_val = ""
        apellidos_val = ""
        for col_idx, header in enumerate(raw_headers):
            h_lower = header.strip().lower()
            val = str(row[col_idx]).strip() if col_idx < len(row) and row[col_idx] is not None else ""
            if h_lower == "email":
                email_val = val
            elif h_lower == "nombre":
                nombre_val = val
            elif h_lower == "apellidos":
                apellidos_val = val

        if not email_val:
            errors.append(f"Fila {row_idx}: email vacío, se omite")
            continue

        valores = {}
        for act in activity_columns:
            col_idx = act["col_idx"]
            val = row[col_idx] if col_idx < len(row) else None
            valores[act["nombre"]] = val

        entries.append({
            "email": email_val,
            "nombre": nombre_val,
            "apellidos": apellidos_val,
            "valores": valores,
        })

    actividades_response = [
        {"nombre": a["nombre"], "tipo": a["tipo"], "valores_muestra": a["valores_muestra"]}
        for a in activity_columns
    ]

    return entries, actividades_response, errors


def _parse_csv(content: bytes) -> tuple[list[dict], list[dict], list[str]]:
    text = content.decode("utf-8-sig")
    lines = text.splitlines()
    if not lines:
        return [], [], ["Archivo vacío"]

    first_line = lines[0]
    delimiter = _detect_csv_delimiter(first_line)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    if reader.fieldnames:
        reader.fieldnames = [f.strip().lower() if f else f for f in reader.fieldnames]

    raw_headers = reader.fieldnames or []
    activity_columns = []
    entries = []
    errors = []

    all_rows = list(reader)
    for col_name in raw_headers:
        sample_values = []
        for row in all_rows[:10]:
            val = row.get(col_name, "")
            if val and val.strip():
                try:
                    sample_values.append(float(val))
                except ValueError:
                    sample_values.append(val)
        col_info = _classify_column(col_name, sample_values)
        if col_info:
            activity_columns.append({**col_info, "col_name": col_name})

    for row_idx, row in enumerate(all_rows, start=2):
        if all(not (row.get(k) or "").strip() for k in METADATA_COLUMNS if row.get(k)):
            if not any(row.get(a["col_name"]) for a in activity_columns):
                continue

        email_val = (row.get("email") or "").strip()
        nombre_val = (row.get("nombre") or "").strip()
        apellidos_val = (row.get("apellidos") or "").strip()

        if not email_val:
            errors.append(f"Fila {row_idx}: email vacío, se omite")
            continue

        valores = {}
        for act in activity_columns:
            val = row.get(act["col_name"], "")
            if val and val.strip():
                try:
                    valores[act["nombre"]] = float(val)
                except ValueError:
                    valores[act["nombre"]] = val.strip()
            else:
                valores[act["nombre"]] = None

        entries.append({
            "email": email_val,
            "nombre": nombre_val,
            "apellidos": apellidos_val,
            "valores": valores,
        })

    actividades_response = [
        {"nombre": a["nombre"], "tipo": a["tipo"], "valores_muestra": a["valores_muestra"]}
        for a in activity_columns
    ]

    return entries, actividades_response, errors


def _parse_finalizacion_xlsx(content: bytes) -> tuple[list[dict], list[str]]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return [], ["Archivo vacío"]

    raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    activity_columns = []
    entries = []
    errors = []

    for col_idx in range(len(raw_headers)):
        cleaned = raw_headers[col_idx].strip()
        if _is_metadata_column(cleaned):
            continue
        if cleaned.endswith(NUMERIC_SUFFIX) or cleaned.endswith(f" {NUMERIC_SUFFIX}"):
            continue  # RN-08: skip numeric
        sample_values = []
        for row in rows[1:11]:
            if col_idx < len(row) and row[col_idx] is not None:
                sample_values.append(str(row[col_idx]))
        text_values = [v for v in sample_values if v in {"Finalizado", "No finalizado", "Entregado", "Pendiente"}]
        if text_values:
            activity_columns.append({"nombre": cleaned, "col_idx": col_idx})

    for row_idx, row in enumerate(rows[1:], start=2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        email_val = ""
        nombre_val = ""
        apellidos_val = ""
        for col_idx, header in enumerate(raw_headers):
            h_lower = header.strip().lower()
            val = str(row[col_idx]).strip() if col_idx < len(row) and row[col_idx] is not None else ""
            if h_lower == "email":
                email_val = val
            elif h_lower == "nombre":
                nombre_val = val
            elif h_lower == "apellidos":
                apellidos_val = val

        if not email_val:
            continue

        estados = {}
        for act in activity_columns:
            col_idx = act["col_idx"]
            val = str(row[col_idx]).strip() if col_idx < len(row) and row[col_idx] is not None else ""
            estados[act["nombre"]] = val

        entries.append({
            "email": email_val,
            "nombre": nombre_val,
            "apellidos": apellidos_val,
            "estados": estados,
        })

    return entries, errors


def _parse_finalizacion_csv(content: bytes) -> tuple[list[dict], list[str]]:
    text = content.decode("utf-8-sig")
    lines = text.splitlines()
    if not lines:
        return [], ["Archivo vacío"]

    first_line = lines[0]
    delimiter = _detect_csv_delimiter(first_line)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    if reader.fieldnames:
        reader.fieldnames = [f.strip().lower() if f else f for f in reader.fieldnames]

    raw_headers = reader.fieldnames or []
    activity_columns = []
    entries = []
    errors = []

    all_rows = list(reader)
    for col_name in raw_headers:
        cleaned = col_name.strip()
        if _is_metadata_column(cleaned):
            continue
        if cleaned.endswith(NUMERIC_SUFFIX) or cleaned.endswith(f" {NUMERIC_SUFFIX}"):
            continue
        sample_values = [row.get(col_name, "").strip() for row in all_rows[:10] if row.get(col_name, "").strip()]
        if any(v in {"Finalizado", "No finalizado", "Entregado", "Pendiente"} for v in sample_values):
            activity_columns.append({"nombre": cleaned, "col_name": col_name})

    for row_idx, row in enumerate(all_rows, start=2):
        email_val = (row.get("email") or "").strip()
        nombre_val = (row.get("nombre") or "").strip()
        apellidos_val = (row.get("apellidos") or "").strip()

        if not email_val:
            continue

        estados = {}
        for act in activity_columns:
            val = row.get(act["col_name"], "").strip()
            estados[act["nombre"]] = val

        entries.append({
            "email": email_val,
            "nombre": nombre_val,
            "apellidos": apellidos_val,
            "estados": estados,
        })

    return entries, errors


class CalificacionesService:

    @staticmethod
    async def preview(materia_id: uuid.UUID, file: UploadFile) -> dict:
        ext = os.path.splitext(file.filename or "archivo")[1].lower()
        if ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Formato no soportado: {ext}. Formatos aceptados: {', '.join(ALLOWED_EXTENSIONS)}",
            )

        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Archivo demasiado grande. Máximo {MAX_FILE_SIZE // (1024*1024)}MB",
            )

        if ext == ".xlsx":
            entries, actividades, errors = _parse_xlsx(content)
        else:
            entries, actividades, errors = _parse_csv(content)

        token = str(uuid.uuid4())
        _preview_store[token] = {
            "entries": entries,
            "materia_id": str(materia_id),
            "actividades": actividades,
            "tipo": "calificaciones",
            "timestamp": time.time(),
        }

        return {
            "preview_token": token,
            "actividades_detectadas": actividades,
            "alumnos_count": len(entries),
            "errores": errors,
        }

    @staticmethod
    async def _get_entradas_by_email(
        db: AsyncSession, materia_id: uuid.UUID, tenant_id: uuid.UUID,
    ) -> dict[str, tuple[uuid.UUID, str, str]]:
        query = select(EntradaPadron).where(
            EntradaPadron.tenant_id == tenant_id,
            EntradaPadron.deleted_at.is_(None),
        )
        result = await db.execute(query)
        entradas = list(result.scalars().all())
        result_map = {}
        for e in entradas:
            try:
                result_map[e.email] = (e.id, e.nombre, e.apellidos)
            except Exception:
                pass
        return result_map

    @staticmethod
    async def derivar_aprobado(
        db: AsyncSession,
        nota_numerica: float | None,
        nota_textual: str | None,
        materia_id: uuid.UUID,
        asignacion_id: uuid.UUID,
        tenant_id: uuid.UUID,
    ) -> bool:
        umbral = None
        if asignacion_id:
            repo = UmbralMateriaRepository(db)
            umbral = await repo.get_by_asignacion_y_materia(asignacion_id, materia_id, tenant_id)

        umbral_pct = umbral.umbral_pct if umbral else DEFAULT_UMBRAL_PCT
        valores_aprobatorios = umbral.valores_aprobatorios if umbral else DEFAULT_VALORES_APROBATORIOS

        if nota_numerica is not None:
            return float(nota_numerica) >= umbral_pct
        elif nota_textual is not None:
            return nota_textual in valores_aprobatorios
        else:
            return False

    @staticmethod
    async def confirm(
        db: AsyncSession,
        preview_token: str,
        user: User,
        actividades_seleccionadas: list[str],
    ) -> dict:
        data = _preview_store.pop(preview_token, None)
        if not data:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Token de preview inválido o expirado",
            )

        if time.time() - data["timestamp"] > PREVIEW_TTL:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Token de preview expirado",
            )

        materia_id = uuid.UUID(data["materia_id"])
        repo_calif = CalificacionesRepository(db)

        email_map = await CalificacionesService._get_entradas_by_email(db, materia_id, user.tenant_id)

        umbral_repo = UmbralMateriaRepository(db)
        asignacion_id = await UmbralMateriaRepository.get_asignacion_id(
            db, user.id, materia_id, user.tenant_id,
        )

        calificaciones_data = []
        errores_matching = []

        for entry in data["entries"]:
            entrada_info = email_map.get(entry["email"])
            if not entrada_info:
                errores_matching.append(f"Alumno {entry['email']} no encontrado en padrón")
                continue

            entrada_id = entrada_info[0]
            for act_nombre in actividades_seleccionadas:
                if act_nombre not in entry["valores"]:
                    continue

                valor = entry["valores"][act_nombre]
                nota_numerica = None
                nota_textual = None

                actividad_info = None
                for a in data["actividades"]:
                    if a["nombre"] == act_nombre:
                        actividad_info = a
                        break

                if actividad_info and actividad_info["tipo"] == "numerica":
                    if valor is not None:
                        try:
                            nota_numerica = float(valor)
                        except (ValueError, TypeError):
                            pass
                else:
                    if valor is not None and str(valor).strip():
                        nota_textual = str(valor).strip()

                if nota_numerica is None and nota_textual is None:
                    continue

                aprobado = await CalificacionesService.derivar_aprobado(
                    db, nota_numerica, nota_textual, materia_id, asignacion_id, user.tenant_id,
                )

                calificaciones_data.append({
                    "entrada_padron_id": entrada_id,
                    "materia_id": materia_id,
                    "actividad": act_nombre,
                    "nota_numerica": nota_numerica,
                    "nota_textual": nota_textual,
                    "aprobado": aprobado,
                })

        created = await repo_calif.bulk_insert(calificaciones_data, user.tenant_id, importado_por=user.id)
        await db.commit()

        aprobados = sum(1 for c in calificaciones_data if c["aprobado"])
        no_aprobados = sum(1 for c in calificaciones_data if not c["aprobado"])

        await AuditService.log_action(
            db=db,
            action="CALIFICACIONES_IMPORTAR",
            user_id=str(user.id),
            resource="calificaciones",
            status="success",
            actor_id=str(user.id),
            materia_id=str(materia_id),
            detalle={
                "tipo": "importacion",
                "actividades": actividades_seleccionadas,
                "errores_matching": len(errores_matching),
            },
            filas_afectadas=len(calificaciones_data),
        )

        return {
            "calificaciones_count": len(calificaciones_data),
            "aprobados_count": aprobados,
            "no_aprobados_count": no_aprobados,
        }

    @staticmethod
    async def preview_finalizacion(
        materia_id: uuid.UUID, file: UploadFile, db: AsyncSession, user: User,
    ) -> dict:
        ext = os.path.splitext(file.filename or "archivo")[1].lower()
        if ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Formato no soportado: {ext}. Formatos aceptados: {', '.join(ALLOWED_EXTENSIONS)}",
            )

        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Archivo demasiado grande. Máximo {MAX_FILE_SIZE // (1024*1024)}MB",
            )

        if ext == ".xlsx":
            entries, errors = _parse_finalizacion_xlsx(content)
        else:
            entries, errors = _parse_finalizacion_csv(content)

        email_map = await CalificacionesService._get_entradas_by_email(db, materia_id, user.tenant_id)
        repo_calif = CalificacionesRepository(db)

        posibles_sin_corregir = []
        for entry in entries:
            entrada_info = email_map.get(entry["email"])
            if not entrada_info:
                continue
            entrada_id = entrada_info[0]

            for act_nombre, estado in entry["estados"].items():
                if estado not in ("Finalizado", "Entregado"):
                    continue

                existing = await repo_calif.get_by_entrada_y_actividad(
                    entrada_id, act_nombre, user.tenant_id,
                )
                if existing is not None:
                    continue

                posibles_sin_corregir.append({
                    "alumno": f"{entry['nombre']} {entry['apellidos']}".strip(),
                    "actividad": act_nombre,
                    "fecha_entrega": None,
                })

        token = str(uuid.uuid4())
        _preview_store[token] = {
            "posibles": posibles_sin_corregir,
            "materia_id": str(materia_id),
            "tipo": "finalizacion",
            "timestamp": time.time(),
        }

        return {
            "preview_token": token,
            "posibles_sin_corregir": posibles_sin_corregir,
        }

    @staticmethod
    async def confirm_finalizacion(
        db: AsyncSession, preview_token: str, user: User,
    ) -> dict:
        data = _preview_store.pop(preview_token, None)
        if not data:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Token de preview inválido o expirado",
            )

        if time.time() - data["timestamp"] > PREVIEW_TTL:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Token de preview expirado",
            )

        materia_id = uuid.UUID(data["materia_id"])
        registros = len(data["posibles"])

        await AuditService.log_action(
            db=db,
            action="CALIFICACIONES_IMPORTAR",
            user_id=str(user.id),
            resource="calificaciones",
            status="success",
            actor_id=str(user.id),
            materia_id=str(materia_id),
            detalle={
                "tipo": "finalizacion",
                "registros_detectados": registros,
            },
            filas_afectadas=0,
        )

        return {"registros_detectados": registros}

    @staticmethod
    async def vaciar_datos(
        db: AsyncSession, materia_id: uuid.UUID, user: User,
    ) -> dict:
        repo = CalificacionesRepository(db)
        eliminadas = await repo.vaciar_datos_usuario(materia_id, user.id, user.tenant_id)
        await db.commit()

        if eliminadas > 0:
            await AuditService.log_action(
                db=db,
                action="CALIFICACIONES_IMPORTAR",
                user_id=str(user.id),
                resource="calificaciones",
                status="success",
                actor_id=str(user.id),
                materia_id=str(materia_id),
                detalle={
                    "tipo": "vaciado",
                    "materia_id": str(materia_id),
                },
                filas_afectadas=eliminadas,
            )

        return {"eliminados_count": eliminadas}
