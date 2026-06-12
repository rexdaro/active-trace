import pytest
import pytest_asyncio
import uuid
import io
import os
from datetime import datetime, timezone
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine, async_sessionmaker
from sqlalchemy import select
from app.models.base import Base
from app.models.tenant import Tenant
from app.models.user import Usuario, User
from app.models.materia import Materia
from app.models.cohorte import Cohorte
from app.models.padron import VersionPadron, EntradaPadron
from app.models.audit import AuditLog
from app.repositories.padron import PadronRepository
from app.services.padron import PadronService, _preview_store
from app.schemas.padron import PadronPreviewResponse

os.environ["ENCRYPTION_KEY"] = "test-key-32-chars-long-for-encryption!!"


@pytest_asyncio.fixture
async def db_session():
    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    session_factory = async_sessionmaker(engine, expire_on_commit=False)
    async with session_factory() as session:
        yield session


@pytest_asyncio.fixture
async def test_tenant(db_session):
    tenant = Tenant(id=uuid.uuid4(), name="Test Tenant")
    db_session.add(tenant)
    await db_session.commit()
    return tenant


@pytest_asyncio.fixture
async def mock_user(db_session, test_tenant):
    user = User(
        id=uuid.uuid4(),
        tenant_id=test_tenant.id,
        email="user@test.com",
        hashed_password="hashed",
        is_2fa_enabled=False,
    )
    db_session.add(user)
    await db_session.commit()
    return user


@pytest_asyncio.fixture
async def test_materia(db_session, test_tenant):
    materia = Materia(id=uuid.uuid4(), tenant_id=test_tenant.id, name="Matemática", code="MAT101", is_active=True)
    db_session.add(materia)
    await db_session.commit()
    return materia


@pytest_asyncio.fixture
async def test_cohorte(db_session, test_tenant):
    cohorte = Cohorte(id=uuid.uuid4(), tenant_id=test_tenant.id, name="2025", carrera_id=uuid.uuid4(), is_active=True)
    db_session.add(cohorte)
    await db_session.commit()
    return cohorte


# ─── Task 15: Test versionado ────────────────────────────────────────────────

class TestVersionado:

    @pytest.mark.asyncio
    async def test_activar_nueva_desactiva_anterior(self, db_session, mock_user, test_tenant, test_materia, test_cohorte):
        repo = PadronRepository(db_session)
        v1 = await repo.crear_version(
            tenant_id=test_tenant.id,
            materia_id=test_materia.id,
            cohorte_id=test_cohorte.id,
            archivo_nombre="test.csv",
            archivo_hash="abc123",
            origen="Archivo",
            cargado_por=mock_user.id,
            activa=True,
        )
        await repo.desactivar_anterior(test_materia.id, test_cohorte.id, test_tenant.id)
        v2 = await repo.crear_version(
            tenant_id=test_tenant.id,
            materia_id=test_materia.id,
            cohorte_id=test_cohorte.id,
            archivo_nombre="test2.csv",
            archivo_hash="def456",
            origen="Archivo",
            cargado_por=mock_user.id,
            activa=True,
        )

        v1_actual = await db_session.get(VersionPadron, v1.id)
        v2_actual = await db_session.get(VersionPadron, v2.id)
        assert v1_actual.activa is False
        assert v2_actual.activa is True

    @pytest.mark.asyncio
    async def test_nueva_version_no_afecta_otras_materias(self, db_session, mock_user, test_tenant, test_materia, test_cohorte):
        repo = PadronRepository(db_session)
        await repo.crear_version(
            tenant_id=test_tenant.id,
            materia_id=test_materia.id,
            cohorte_id=test_cohorte.id,
            archivo_nombre="v1.csv",
            archivo_hash="aaa",
            origen="Archivo",
            cargado_por=mock_user.id,
            activa=True,
        )
        otra_materia = Materia(id=uuid.uuid4(), tenant_id=test_tenant.id, name="Física", code="FIS101", is_active=True)
        db_session.add(otra_materia)
        await db_session.commit()

        await repo.desactivar_anterior(test_materia.id, test_cohorte.id, test_tenant.id)
        v2 = await repo.crear_version(
            tenant_id=test_tenant.id,
            materia_id=test_materia.id,
            cohorte_id=test_cohorte.id,
            archivo_nombre="v2.csv",
            archivo_hash="bbb",
            origen="Archivo",
            cargado_por=mock_user.id,
            activa=True,
        )

        query = select(VersionPadron).where(
            VersionPadron.materia_id == otra_materia.id,
            VersionPadron.activa == True,
        )
        result = await db_session.execute(query)
        otras_activas = list(result.scalars().all())
        assert len(otras_activas) == 0


# ─── Tests de import xlsx y csv (Task 16-17) ────────────────────────────────

class TestImportXlsx:

    @pytest.mark.asyncio
    async def test_preview_xlsx_exitoso(self, db_session, mock_user, test_tenant, test_materia, test_cohorte):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["nombre", "apellidos", "email", "comision", "regional"])
        ws.append(["Juan", "Pérez", "juan@test.com", "A", "CABA"])
        ws.append(["María", "García", "maria@test.com", "B", "GBA"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        class FakeUploadFile:
            filename = "test.xlsx"
            async def read(self):
                return buf.getvalue()

        result = await PadronService.preview(FakeUploadFile(), test_materia.id, test_cohorte.id)

        assert "preview_token" in result
        assert result["filas_count"] == 2
        assert "nombre" in result["columnas_detectadas"]
        assert len(result["errores"]) == 0

    @pytest.mark.asyncio
    async def test_confirm_xlsx_crea_entradas(self, db_session, mock_user, test_tenant, test_materia, test_cohorte):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["nombre", "apellidos", "email", "comision", "regional"])
        ws.append(["Juan", "Pérez", "juan@test.com", "A", "CABA"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        class FakeUploadFile:
            filename = "test.xlsx"
            async def read(self):
                return buf.getvalue()

        preview = await PadronService.preview(FakeUploadFile(), test_materia.id, test_cohorte.id)
        token = preview["preview_token"]

        confirm = await PadronService.confirm(db_session, token, mock_user)

        assert "version_id" in confirm
        assert confirm["entradas_count"] == 1

    @pytest.mark.asyncio
    async def test_confirm_audit_logged(self, db_session, mock_user, test_tenant, test_materia, test_cohorte):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["nombre", "apellidos", "email"])
        ws.append(["Ana", "López", "ana@test.com"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        class FakeUploadFile:
            filename = "test.xlsx"
            async def read(self):
                return buf.getvalue()

        preview = await PadronService.preview(FakeUploadFile(), test_materia.id, test_cohorte.id)
        await PadronService.confirm(db_session, preview["preview_token"], mock_user)

        stmt = select(AuditLog).where(AuditLog.action == "PADRON_CARGAR")
        result = await db_session.execute(stmt)
        logs = list(result.scalars().all())
        assert len(logs) == 1
        assert logs[0].filas_afectadas == 1


class TestImportCsv:

    @pytest.mark.asyncio
    async def test_preview_csv_exitoso(self, db_session, mock_user, test_tenant, test_materia, test_cohorte):
        csv_content = "nombre,apellidos,email,comision,regional\nCarlos,Gomez,carlos@test.com,C,Rosario\n"
        class FakeUploadFile:
            filename = "test.csv"
            async def read(self):
                return csv_content.encode("utf-8")

        result = await PadronService.preview(FakeUploadFile(), test_materia.id, test_cohorte.id)

        assert result["filas_count"] == 1
        assert "email" in result["columnas_detectadas"]

    @pytest.mark.asyncio
    async def test_csv_con_delimitador_punto_coma(self, db_session, mock_user, test_tenant, test_materia, test_cohorte):
        csv_content = "nombre;apellidos;email\nPedro;Martínez;pedro@test.com\n"
        class FakeUploadFile:
            filename = "test.csv"
            async def read(self):
                return csv_content.encode("utf-8")

        result = await PadronService.preview(FakeUploadFile(), test_materia.id, test_cohorte.id)

        assert result["filas_count"] == 1
        assert result["errores"] == []

    @pytest.mark.asyncio
    async def test_csv_confirm_crea_entradas(self, db_session, mock_user, test_tenant, test_materia, test_cohorte):
        csv_content = "nombre,apellidos,email\nLucía,Fernández,lucia@test.com\n"
        class FakeUploadFile:
            filename = "test.csv"
            async def read(self):
                return csv_content.encode("utf-8")

        preview = await PadronService.preview(FakeUploadFile(), test_materia.id, test_cohorte.id)
        confirm = await PadronService.confirm(db_session, preview["preview_token"], mock_user)

        assert confirm["entradas_count"] == 1


# ─── Task 18: Entrada sin usuario_id ──────────────────────────────────────────

class TestEntradaSinUsuario:

    @pytest.mark.asyncio
    async def test_alumno_sin_cuenta(self, db_session, mock_user, test_tenant, test_materia, test_cohorte):
        csv_content = "nombre,apellidos,email\nNuevo,Alumno,noexiste@test.com\n"
        class FakeUploadFile:
            filename = "test.csv"
            async def read(self):
                return csv_content.encode("utf-8")

        preview = await PadronService.preview(FakeUploadFile(), test_materia.id, test_cohorte.id)
        confirm = await PadronService.confirm(db_session, preview["preview_token"], mock_user)

        stmt = select(EntradaPadron).where(EntradaPadron.version_id == confirm["version_id"])
        result = await db_session.execute(stmt)
        entrada = result.scalar_one()
        assert entrada.usuario_id is None
        assert entrada.nombre == "Nuevo"
        assert entrada.apellidos == "Alumno"

    @pytest.mark.asyncio
    async def test_alumno_con_cuenta_existente(self, db_session, mock_user, test_tenant, test_materia, test_cohorte):
        usuario = Usuario(
            id=mock_user.id,
            tenant_id=test_tenant.id,
        )
        usuario.dni = "12345678"
        usuario.cuil = "20123456782"
        usuario.email = "existente@test.com"
        db_session.add(usuario)
        await db_session.commit()
        await db_session.refresh(usuario)

        csv_content = f"nombre,apellidos,email\nAlumno,Existente,existente@test.com\n"
        class FakeUploadFile:
            filename = "test.csv"
            async def read(self):
                return csv_content.encode("utf-8")

        preview = await PadronService.preview(FakeUploadFile(), test_materia.id, test_cohorte.id)
        confirm = await PadronService.confirm(db_session, preview["preview_token"], mock_user)

        stmt = select(EntradaPadron).where(EntradaPadron.version_id == confirm["version_id"])
        result = await db_session.execute(stmt)
        entrada = result.scalar_one()
        assert entrada.usuario_id == usuario.id


# ─── Task 19: Tenant isolation ──────────────────────────────────────────────

class TestTenantIsolation:

    @pytest.mark.asyncio
    async def test_datos_aislados_por_tenant(self, db_session, mock_user, test_tenant, test_materia, test_cohorte):
        repo = PadronRepository(db_session)
        await repo.crear_version(
            tenant_id=test_tenant.id,
            materia_id=test_materia.id,
            cohorte_id=test_cohorte.id,
            archivo_nombre="v1.csv",
            archivo_hash="h1",
            origen="Archivo",
            cargado_por=mock_user.id,
            activa=True,
        )

        otro_tenant_id = uuid.uuid4()
        versiones_otro = await repo.get_versiones_materia(test_materia.id, otro_tenant_id)
        assert len(versiones_otro) == 0

    @pytest.mark.asyncio
    async def test_insercion_protege_tenant_id(self, db_session, mock_user, test_tenant, test_materia, test_cohorte):
        repo = PadronRepository(db_session)
        otro_tenant_id = uuid.uuid4()
        v = await repo.crear_version(
            tenant_id=otro_tenant_id,
            materia_id=test_materia.id,
            cohorte_id=test_cohorte.id,
            archivo_nombre="v1.csv",
            archivo_hash="h1",
            origen="Archivo",
            cargado_por=mock_user.id,
            activa=True,
        )
        assert v.tenant_id == otro_tenant_id

        query = select(VersionPadron).where(
            VersionPadron.tenant_id == test_tenant.id,
            VersionPadron.materia_id == test_materia.id,
        )
        result = await db_session.execute(query)
        assert len(list(result.scalars().all())) == 0


# ─── Task 22: Vaciado RN-04 ─────────────────────────────────────────────────

class TestVaciadoRN04:

    @pytest.mark.asyncio
    async def test_vaciar_solo_propios_datos(self, db_session, test_tenant, test_materia, test_cohorte):
        user1 = User(id=uuid.uuid4(), tenant_id=test_tenant.id, email="u1@test.com",
                      hashed_password="h", is_2fa_enabled=False)
        user2 = User(id=uuid.uuid4(), tenant_id=test_tenant.id, email="u2@test.com",
                      hashed_password="h", is_2fa_enabled=False)
        db_session.add_all([user1, user2])
        await db_session.commit()

        repo = PadronRepository(db_session)
        v1 = await repo.crear_version(
            tenant_id=test_tenant.id, materia_id=test_materia.id, cohorte_id=test_cohorte.id,
            archivo_nombre="u1.csv", archivo_hash="h1", origen="Archivo",
            cargado_por=user1.id, activa=True,
        )
        v2 = await repo.crear_version(
            tenant_id=test_tenant.id, materia_id=test_materia.id, cohorte_id=test_cohorte.id,
            archivo_nombre="u2.csv", archivo_hash="h2", origen="Archivo",
            cargado_por=user2.id, activa=True,
        )

        eliminadas = await repo.vaciar_datos_usuario(test_materia.id, user1.id, test_tenant.id)
        assert eliminadas == 1

        v2_db = await db_session.get(VersionPadron, v2.id)
        assert v2_db is not None
        assert v2_db.cargado_por == user2.id

    @pytest.mark.asyncio
    async def test_vaciar_sin_datos_propios(self, db_session, test_tenant, test_materia, test_cohorte):
        user = User(id=uuid.uuid4(), tenant_id=test_tenant.id, email="u@test.com",
                     hashed_password="h", is_2fa_enabled=False)
        db_session.add(user)
        await db_session.commit()

        repo = PadronRepository(db_session)
        await repo.crear_version(
            tenant_id=test_tenant.id, materia_id=test_materia.id, cohorte_id=test_cohorte.id,
            archivo_nombre="other.csv", archivo_hash="h1", origen="Archivo",
            cargado_por=uuid.uuid4(), activa=True,
        )

        eliminadas = await repo.vaciar_datos_usuario(test_materia.id, user.id, test_tenant.id)
        assert eliminadas == 0

    @pytest.mark.asyncio
    async def test_vaciar_servicio_audita(self, db_session, mock_user, test_tenant, test_materia, test_cohorte):
        repo = PadronRepository(db_session)
        await repo.crear_version(
            tenant_id=test_tenant.id, materia_id=test_materia.id, cohorte_id=test_cohorte.id,
            archivo_nombre="v1.csv", archivo_hash="h1", origen="Archivo",
            cargado_por=mock_user.id, activa=True,
        )

        result = await PadronService.vaciar_datos(db_session, test_materia.id, mock_user)
        assert result["eliminadas"] == 1

        stmt = select(AuditLog).where(AuditLog.action == "PADRON_CARGAR")
        logs = (await db_session.execute(stmt)).scalars().all()
        assert len(logs) == 1
